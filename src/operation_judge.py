# @author JL
# @date 2023/9/25 16:41
import openpyxl
from openpyxl import load_workbook
from pyproj import Transformer
from pyproj import CRS
import math


def transform_coordinates(lon, lat):
    """
    坐标转换函数
    :param lon:     经度
    :param lat:     纬度
    :return:    返回平面坐标
    """
    crs = CRS.from_epsg(4326)  # 大地坐标系（先纬度 再经度）
    crs_cs = CRS.from_epsg(3857)  # 墨卡托投影坐标系

    transformer = Transformer.from_crs(crs, crs_cs)
    x, y = transformer.transform(lat, lon)
    return x, y


def write_excel_xlsx(path, sheet_name, value):
    """
    将坐标转换后的点集写回excel
    :param path:        文件存储路径
    :param sheet_name:  工作表名
    :param value:       需存储的数据
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(len(value)):
        for j in range(len(value[0])):
            sheet.cell(row=i + 1, column=j + 1, value=value[i][j])
    workbook.save(path)
    print("数据已存储在: ", path)


def judge_points(file_name):
    """
    散乱点集判别
    :param file_name: 文件存储路径
    """
    # 读取工作表
    workbook = load_workbook(filename=file_name, data_only=True)
    sheet = workbook.active
    print("数据表尺寸为：", sheet.dimensions)

    cell = sheet[sheet.dimensions]  # 确定处理数据范围
    points = {}  # 存储点集平面坐标信息
    index = 0

    for point in cell:

        lon, lat = point[5].value, point[6].value  # 确保经度、纬度信息位于 F G 列
        x, y = transform_coordinates(lon, lat)

        if len(points) == 0 or [x, y] != points[len(points) - 1]:   # 点集去重复
            points[index] = [x, y]
            index += 1

    print("点集长度为：", len(points))

    points_angle = {}  # 存储点集角度信息，第 1 个点没有瞬时角度

    for i in range(1, len(points)):

        point = points[i]
        point_prev = points[i - 1]
        angle = math.degrees(math.atan2(point[1] - point_prev[1], point[0] - point_prev[0]))
        points_angle[i] = angle

    # print("有效点集的角度值：", points_angle)

    # 以 11 个点为窗口大小的角度总和
    angle_sum_11 = 0
    for i in range(1, 12):
        angle_sum_11 += points_angle[i]     # 滑动窗口初始值

    points_bias = {}  # 瞬时角度偏差
    for i in range(6, len(points_angle) - 4):
        angle_c = points_angle[i]   # 从第 7 个点开始计算瞬时角度偏差(因为第 1 个点没有瞬时角度)
        if i != 6:
            angle_sum_11 -= points_angle[i - 5]
            angle_sum_11 += points_angle[i + 5]
        angle_bias = (angle_sum_11 / 11) - angle_c

        points_bias[i] = angle_bias

    # print("角度瞬时偏差：", points_bias)

    point_bias_threshold = 15  # 角度偏差阈值
    point_num_threshold = 15  # 连续满足阈值条件的点数

    isOperation = False
    section_bias_threshold = 5     # 区间角度差异阈值
    section_angle_mean_prev = 1000  # 存储区间范围内的角度均值

    begin_index = -1

    #  遍历查找满足条件的连续区间
    for i in range(6, len(points_angle) - 4):
        if points_bias[i] <= point_bias_threshold and begin_index == -1:
            begin_index = i
        elif points_bias[i] > point_bias_threshold:
            # 找到连续的N个点满足条件时，计算该区间的角度均值
            if begin_index != -1 and i - begin_index > point_num_threshold:
                section_angle_sum = 0
                for index in range(begin_index, i):
                    section_angle_sum += points_angle[index]
                section_angle_mean = section_angle_sum / (i - begin_index)

                # 如果此前已有满足条件的连续区间，并且两区间角度均值小于阈值，则直接判定该点集为作业点集，退出循环
                if section_angle_mean_prev != 1000 and math.fabs(section_angle_mean_prev - section_angle_mean) < section_bias_threshold:
                    isOperation = True
                    break

                # 将当前区间的角度均值赋值给prev，用于后续区间比较
                section_angle_mean_prev = section_angle_mean

            begin_index = -1    # 当前点不满足角度偏差阈值，区间下标初始化

    if isOperation:
        print("正常作业点集")
    else:
        print("散乱点集")


if __name__ == "__main__":
    judge_points("../data/origin/machine_operation_12.xlsx")
